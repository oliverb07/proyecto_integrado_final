from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from .models import Bitacora, Usuario
from neonatos.models import Madre, Parto, RecienNacido
from datetime import datetime
from django.utils import timezone
from decimal import Decimal
import io


# Página de inicio
def inicio(request):
    return render(request, 'base.html')

# Página del generador de reportes
def vistaReportes(request):
    return render(request, 'GeneradorReporte/reportes.html')

def vistaReportea09(request):
    return render(request, 'GeneradorReporte/reporte_a09.html')

def vistaReportea04(request):
    return render(request, 'GeneradorReporte/reporte_a04.html')

# Vista de Bitácora (solo supervisores)

def verBitacora(request):
    # Obtener todos los registros de la bitácora ordenados por fecha descendente
    logs = Bitacora.objects.select_related('usuario').order_by('-fecha_hora')
    return render(request, 'GeneradorReporte/bitacora.html', {'logs': logs})


def split_rut_dv(rut_normalizado: str):
    """Espera rut sin puntos y con guion o ya normalizado '12345678-9' o '123456789'"""
    if "-" in rut_normalizado:
        parts = rut_normalizado.split("-")
        return parts[0], parts[1]
    # si no tiene guion, asumir último carácter es DV
    if len(rut_normalizado) > 1:
        return rut_normalizado[:-1], rut_normalizado[-1]
    return rut_normalizado, ""

# Exportación de Excel
def robson_group_for_parto(parto: Parto):
    """
    Clasificación simplificada Robson basada en tu modelo:
    Reglas aplicadas (resumen):
      - paridad: madre.paridad ('nulipara' / 'multipara')
      - cesareas_previas: madre.cesareas_previas (int)
      - embarazo_multiple: parto.embarazo_multiple (bool)
      - presentacion_fetal: parto.presentacion_fetal ('cefalica','pelvica','transversa')
      - edad_gestacional: parto.edad_gestacional (semanas)
      - tipo_parto: parto.tipo_parto (cesarea_... / vaginal / instrumental)
    Devuelve grupo int 1..10 o None si no clasifica.
    """
    madre = parto.madre
    paridad = getattr(madre, "paridad", "nulipara")
    ces_prev = getattr(madre, "cesareas_previas", 0)
    multifetal = bool(parto.embarazo_multiple)
    present = parto.presentacion_fetal or "cefalica"
    edad = parto.edad_gestacional or 0
    cesarea = parto.tipo_parto in ("cesarea_electiva", "cesarea_urgencia")

    # Grupo 1: Nulíparas, embarazo único, cefálica, >=37, parto espóntaneo (vaginal)
    if paridad == "nulipara" and not multifetal and present == "cefalica" and edad >= 37 and not cesarea:
        return 1

    # Grupo 2: Nulíparas, único, cefálica, >=37, cesárea programada o inducción -> map si cesárea o inicio_inducido
    if paridad == "nulipara" and not multifetal and present == "cefalica" and edad >= 37:
        # si fue cesárea -> 2, si fue parto vaginal inducido pero terminó vaginal quizá 2.a/2.b en excel original; simplificamos a 2
        return 2

    # Grupo 3: Multípara sin cesárea previa, único, cefálica, >=37, parto espontáneo (no cesárea)
    if paridad == "multipara" and ces_prev == 0 and not multifetal and present == "cefalica" and edad >= 37 and not cesarea:
        return 3

    # Grupo 4: Multípara sin cesárea previa, único, cefálica, >=37, cesárea programada o inducción
    if paridad == "multipara" and ces_prev == 0 and not multifetal and present == "cefalica" and edad >= 37:
        return 4

    # Grupo 5: Multípara con al menos 1 cesárea previa, embarazo único, cefálica, >=37
    if paridad == "multipara" and ces_prev >= 1 and not multifetal and present == "cefalica" and edad >= 37:
        return 5

    # Grupo 6: Nulíparas, único, podálica
    if paridad == "nulipara" and present == "pelvica" and not multifetal:
        return 6

    # Grupo 7: Multíparas, único, podálica (con o sin cesáreas previas)
    if paridad == "multipara" and present == "pelvica" and not multifetal:
        return 7

    # Grupo 8: Embarazo múltiple
    if multifetal:
        return 8

    # Grupo 9: Transversa u oblicua (todas)
    if present == "transversa":
        return 9

    # Grupo 10: Todas las mujeres con embarazo único, cefálica, <37 semanas
    if present == "cefalica" and edad < 37 and not multifetal:
        return 10

    return None

# --- Excel generation --- #

def build_rem_sheet(wb: Workbook, partidas_qs):
    """
    REM: contadores para las filas solicitadas.
    partidas_qs: queryset de Parto
    """
    ws = wb.create_sheet("REM")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---------------------------
    # ENCABEZADO (mantener la estética que pediste)
    # ---------------------------
    ws.merge_cells("A1:A3")
    ws["A1"] = "CARACTERÍSTICAS DEL PARTO"
    ws["A1"].font = bold
    ws["A1"].alignment = center
    ws["A1"].border = border

    ws.merge_cells("B1:B3")
    ws["B1"] = "TOTAL"
    ws["B1"].font = bold
    ws["B1"].alignment = center
    ws["B1"].border = border

    # Partos según edad (C-F)
    ws.merge_cells("C1:F1")
    ws["C1"] = "PARTOS SEGÚN EDAD DE LA MADRE"
    ws["C1"].font = bold
    ws["C1"].alignment = center
    ws["C1"].border = border

    edades = ["<15 AÑOS", "15 A 19 AÑOS", "20 A 34 AÑOS", "≥35 AÑOS"]
    col = 3
    for e in edades:
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        cell = ws.cell(row=2, column=col, value=e)
        cell.font = bold
        cell.alignment = center
        cell.border = border
        col += 1

    # Partos prematuros (G-J)
    ws.merge_cells("G1:J1")
    ws["G1"] = "PARTOS PREMATUROS (>22 semanas)"
    ws["G1"].font = bold
    ws["G1"].alignment = center
    ws["G1"].border = border

    prematuros = [
        "Menos de 24 semanas",
        "24 a 28 semanas",
        "29 a 32 semanas",
        "33 a 36 semanas",
    ]
    for p in prematuros:
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        c = ws.cell(row=2, column=col, value=p)
        c.font = bold
        c.alignment = center
        c.border = border
        col += 1

    # Oxitocina profiláctica (K)
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(row=1, column=col, value="Uso de oxitocina profiláctica").font = bold
    ws.cell(row=1, column=col).alignment = center
    ws.cell(row=1, column=col).border = border
    col += 1

    # Analgesias (L-Q) 6 columnas
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+5)
    ws.cell(row=1, column=col, value="Anestesia y/o Analgesia").font = bold
    ws.cell(row=1, column=col).alignment = center
    ws.cell(row=1, column=col).border = border

    analgesias = [
        ("neuroaxial", "Neuroaxial"),
        ("oxido_nitroso", "Óxido nitroso"),
        ("endovenosa", "Endovenosa"),
        ("general", "General"),
        ("local", "Local"),
        ("no_farmacologica", "No farmacológica"),
    ]
    analg_start_col = col
    for key, label in analgesias:
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        c = ws.cell(row=2, column=col, value=label)
        c.font = bold
        c.alignment = center
        c.border = border
        col += 1

    # Ligadura tardía del cordón (si no hay campo, lo dejamos en 0)
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(row=1, column=col, value="Ligadura tardía del cordón (>60 seg)").font = bold
    ws.cell(row=1, column=col).alignment = center
    ws.cell(row=1, column=col).border = border
    cordon_col = col
    col += 1

    # Contacto piel a piel >30 min (Madre) — ocupará 4 columnas (S..V por ejemplo)
    piel_start = col
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+3)
    ws.cell(row=1, column=col, value="Contacto piel a piel >30 min (Madre)").font = bold
    ws.cell(row=1, column=col).alignment = center
    ws.cell(row=1, column=col).border = border

    # Subencabezados (RN ≤ 2.499g y RN ≥ 2.500g), cada uno ocupará 2 columnas (para estética)
    # Escribimos el texto en la esquina superior izquierda del merge.
    ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col+1)
    ws.cell(row=2, column=col, value="RN peso ≤ 2.499g").font = bold
    ws.cell(row=2, column=col).alignment = center
    ws.cell(row=2, column=col).border = border

    ws.merge_cells(start_row=2, start_column=col+2, end_row=3, end_column=col+3)
    ws.cell(row=2, column=col+2, value="RN peso ≥ 2.500g").font = bold
    ws.cell(row=2, column=col+2).alignment = center
    ws.cell(row=2, column=col+2).border = border

    piel_le_col = col
    piel_ge_col = col + 2
    col += 4

    # Lactancia 60 min
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(row=1, column=col, value="Lactancia primeros 60 min").font = bold
    ws.cell(row=1, column=col).alignment = center
    ws.cell(row=1, column=col).border = border
    lact_col = col
    col += 1

    # Alojamiento conjunto
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(row=1, column=col, value="Alojamiento conjunto").font = bold
    ws.cell(row=1, column=col).alignment = center
    ws.cell(row=1, column=col).border = border
    alojamiento_col = col
    col += 1

    # Pertinencia cultural (campo no existe en tu modelo -> 0)
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(row=1, column=col, value="Atención con pertinencia cultural").font = bold
    ws.cell(row=1, column=col).alignment = center
    ws.cell(row=1, column=col).border = border
    cultural_col = col
    col += 1

    # Pueblos originarios
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(row=1, column=col, value="Pueblos originarios").font = bold
    ws.cell(row=1, column=col).alignment = center
    ws.cell(row=1, column=col).border = border
    pueblos_col = col
    col += 1

    # Migrantes
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(row=1, column=col, value="Migrantes").font = bold
    ws.cell(row=1, column=col).alignment = center
    ws.cell(row=1, column=col).border = border
    migrantes_col = col
    col += 1

    # Discapacidad
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(row=1, column=col, value="Discapacidad").font = bold
    ws.cell(row=1, column=col).alignment = center
    ws.cell(row=1, column=col).border = border
    disc_col = col
    col += 1

    # Privada de libertad
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(row=1, column=col, value="Privada de libertad").font = bold
    ws.cell(row=1, column=col).alignment = center
    ws.cell(row=1, column=col).border = border
    privada_col = col
    col += 1

    # ---------------------------
    # Ahora las filas (tu lista original)
    # ---------------------------
    rows = [
        "TOTAL PARTOS",
        "VAGINAL",
        "INSTRUMENTAL",
        "CESÁREA ELECTIVA",
        "CESÁREA URGENCIA",
        "PARTO PREHOSPITALARIO",
        "Plan de parto",
        "ENTREGA DE PLACENTA A SOLICITUD",
        "EMBARAZO NO CONTROLADO",
        "PARTO EN DOMICILIO - CON ATENCIÓN PROFESIONAL",
        "PARTO EN DOMICILIO - SIN ATENCIÓN PROFESIONAL",
    ]

    start_row = 4

    # Helper: devolver queryset de partos según fila (label)
    def qs_por_fila(label):
        if label == "TOTAL PARTOS":
            return partidas_qs
        if label == "VAGINAL":
            return partidas_qs.filter(tipo_parto="vaginal")
        if label == "INSTRUMENTAL":
            return partidas_qs.filter(tipo_parto="instrumental")
        if label == "CESÁREA ELECTIVA":
            return partidas_qs.filter(tipo_parto="cesarea_electiva")
        if label == "CESÁREA URGENCIA":
            return partidas_qs.filter(tipo_parto="cesarea_urgencia")
        if label == "PARTO PREHOSPITALARIO":
            return partidas_qs.filter(tipo_parto="prehospitalario")
        if label == "Plan de parto":
            return partidas_qs.filter(plan_parto=True)
        if label == "ENTREGA DE PLACENTA A SOLICITUD":
            # no hay campo → devolvemos empty queryset
            return partidas_qs.none()
        if label == "EMBARAZO NO CONTROLADO":
            return partidas_qs.filter(madre__controles_prenatales__iexact="no")
        if label == "PARTO EN DOMICILIO - CON ATENCIÓN PROFESIONAL":
            return partidas_qs.filter(tipo_parto="domicilio", registrado_por__isnull=False)
        if label == "PARTO EN DOMICILIO - SIN ATENCIÓN PROFESIONAL":
            return partidas_qs.filter(tipo_parto="domicilio", registrado_por__isnull=True)
        return partidas_qs.none()

    # Recorremos filas y escribimos datos por columna
    r = start_row
    for label in rows:
        qs = qs_por_fila(label)

        # Etiqueta
        cell_label = ws.cell(row=r, column=1, value=label)
        cell_label.font = bold
        cell_label.alignment = center
        cell_label.border = border

        # TOTAL
        total_cell = ws.cell(row=r, column=2, value=qs.count())
        total_cell.alignment = center
        total_cell.border = border

        # EDADES C-F (columnas 3..6)
        # <15, 15-19, 20-34, >=35
        edades_filters = [
            ("lt", {"madre__edad__lt": 15}),
            ("15_19", {"madre__edad__gte": 15, "madre__edad__lte": 19}),
            ("20_34", {"madre__edad__gte": 20, "madre__edad__lte": 34}),
            ("ge35", {"madre__edad__gte": 35}),
        ]
        col_e = 3
        for name, f in edades_filters:
            cnt = qs.filter(**f).count()
            c = ws.cell(row=r, column=col_e, value=cnt)
            c.alignment = center
            c.border = border
            col_e += 1

        # PREMATUROS G-J (col actual after ages)
        # "Menos de 24 semanas", "24 a 28", "29 a 32", "33 a 36"
        prem_filters = [
            {"edad_gestacional__lt": 24},
            {"edad_gestacional__gte": 24, "edad_gestacional__lte": 28},
            {"edad_gestacional__gte": 29, "edad_gestacional__lte": 32},
            {"edad_gestacional__gte": 33, "edad_gestacional__lte": 36},
        ]
        col_p = 7
        for f in prem_filters:
            cnt = qs.filter(**f).count()
            c = ws.cell(row=r, column=col_p, value=cnt)
            c.alignment = center
            c.border = border
            col_p += 1

        # Oxitocina (col K = 11)
        ox_col = 11
        c = ws.cell(row=r, column=ox_col, value=qs.filter(oxitocina=True).count())
        c.alignment = center
        c.border = border

        # ANALGESIAS L-Q (cols 12..17)
        anal_col = analg_start_col
        for key, label_txt in analgesias:
            cnt = qs.filter(analgesia=key).count()
            c = ws.cell(row=r, column=anal_col, value=cnt)
            c.alignment = center
            c.border = border
            anal_col += 1

        # Ligadura cordón (col cordon_col)
        # Si no hay campo en el model -> dejar 0
        try:
            # si existiera un campo 'ligadura_tardia' en Parto:
            cnt_cordon = qs.filter(ligadura_tardia=True).count()
        except Exception:
            cnt_cordon = 0
        c = ws.cell(row=r, column=cordon_col, value=cnt_cordon)
        c.alignment = center
        c.border = border

        # Contacto piel a piel >30 min (Madre)
        # RN ≤ 2.499g  -> contamos partos que tengan al menos un RN con peso <= 2.499 y contacto_piel_piel True
        cnt_piel_le = qs.filter(contacto_piel_piel=True, recien_nacidos__peso__lte=Decimal("2.499")).distinct().count()
        c = ws.cell(row=r, column=piel_le_col, value=cnt_piel_le)
        c.alignment = center
        c.border = border

        # Espacio extra dentro del merge (col piel_le_col+1) — ponemos la misma cifra para estética
        c2 = ws.cell(row=r, column=piel_le_col+1, value=cnt_piel_le)
        c2.alignment = center
        c2.border = border

        # RN ≥ 2.500g
        cnt_piel_ge = qs.filter(contacto_piel_piel=True, recien_nacidos__peso__gte=Decimal("2.500")).distinct().count()
        c3 = ws.cell(row=r, column=piel_ge_col, value=cnt_piel_ge)
        c3.alignment = center
        c3.border = border
        c4 = ws.cell(row=r, column=piel_ge_col+1, value=cnt_piel_ge)
        c4.alignment = center
        c4.border = border

        # Lactancia primeros 60 min -> campo no existe: 0 (si más tarde lo agregas cambia el filtro)
        ws.cell(row=r, column=lact_col, value=0).alignment = center
        ws.cell(row=r, column=lact_col).border = border

        # Alojamiento conjunto
        ws.cell(row=r, column=alojamiento_col, value=qs.filter(alojamiento_conjunto=True).count()).alignment = center
        ws.cell(row=r, column=alojamiento_col).border = border

        # Pertinencia cultural -> no existe -> 0
        ws.cell(row=r, column=cultural_col, value=0).alignment = center
        ws.cell(row=r, column=cultural_col).border = border

        # Pueblos originarios -> contamos madre__pueblo_originario = "si"
        ws.cell(row=r, column=pueblos_col, value=qs.filter(madre__pueblo_originario__iexact="si").distinct().count()).alignment = center
        ws.cell(row=r, column=pueblos_col).border = border

        # Migrantes -> madre__nacionalidad == "migrante"
        ws.cell(row=r, column=migrantes_col, value=qs.filter(madre__nacionalidad__iexact="migrante").distinct().count()).alignment = center
        ws.cell(row=r, column=migrantes_col).border = border

        # Discapacidad -> madre__discapacidad == "si"
        ws.cell(row=r, column=disc_col, value=qs.filter(madre__discapacidad__iexact="si").distinct().count()).alignment = center
        ws.cell(row=r, column=disc_col).border = border

        # Privada de libertad -> madre__privada_libertad == "si"
        ws.cell(row=r, column=privada_col, value=qs.filter(madre__privada_libertad__iexact="si").distinct().count()).alignment = center
        ws.cell(row=r, column=privada_col).border = border

        r += 1

    # ---------------------------
    # Ajustes de ancho de columna (para que el texto no se corte)
    # ---------------------------
    # Establecemos anchos razonables (ajusta si quieres)
    from openpyxl.utils import get_column_letter

    for i in range(1, col):
        letter = get_column_letter(i)

        if i == 1:
            ws.column_dimensions[letter].width = 40  # La columna A
        else:
            ws.column_dimensions[letter].width = 16  # Las demás


def build_aps_sheet(wb: Workbook, start_date=None, end_date=None):
    """
    APS: una fila por RN (o por Parto si prefieres).
    Campos solicitados: Fecha, Hora, Nombre, RUT, DV, Tipo de parto, Peso, Talla, Apgar1, Apgar5, APEGO (contacto piel a piel)
    Opcional: filtrar por rango de fechas si start_date/end_date se pasan.
    """
    ws = wb.create_sheet("APS")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    headers = ["Fecha", "Hora", "Nombre madre", "RUT", "DV", "Tipo de parto", "Peso (kg)", "Talla (cm)", "Apgar 1", "Apgar 5", "Apego (piel a piel)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = bold
        ws.cell(row=1, column=c).alignment = center

    # Query: todos los RN relacionados a partos en rango (si aplica)
    rns = RecienNacido.objects.select_related("parto__madre", "parto").order_by("id")
    if start_date:
        rns = rns.filter(parto__fecha_parto__gte=start_date)
    if end_date:
        rns = rns.filter(parto__fecha_parto__lte=end_date)

    row = 2
    for rn in rns:
        parto = rn.parto
        madre = parto.madre
        rut_raw = getattr(madre, "rut", "") or ""
        rut_num, dv = split_rut_dv(rut_raw)
        fecha = parto.fecha_parto
        hora = getattr(parto, "hora_parto", None)
        apego = "Sí" if parto.contacto_piel_piel else "No"

        ws.cell(row=row, column=1, value=fecha.strftime("%Y-%m-%d") if fecha else "")
        ws.cell(row=row, column=2, value=hora.strftime("%H:%M") if hora else "")
        ws.cell(row=row, column=3, value=f"{madre.nombres} {madre.apellidos}")
        ws.cell(row=row, column=4, value=rut_num)
        ws.cell(row=row, column=5, value=dv)
        ws.cell(row=row, column=6, value=parto.get_tipo_parto_display() if hasattr(parto, "get_tipo_parto_display") else parto.tipo_parto)
        ws.cell(row=row, column=7, value=float(rn.peso) if rn.peso is not None else "")
        ws.cell(row=row, column=8, value=float(rn.talla) if rn.talla is not None else "")
        ws.cell(row=row, column=9, value=rn.apgar_1 if rn.apgar_1 is not None else "")
        ws.cell(row=row, column=10, value=rn.apgar_5 if rn.apgar_5 is not None else "")
        ws.cell(row=row, column=11, value=apego)
        row += 1

    thin = Side(border_style="thin", color="000000")

    for r in range(1, row):   # desde fila 1 hasta última fila escrita
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )
    
    # Ajustes de ancho
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

def build_robson_sheet(wb: Workbook, partidas_qs):
    """
    ROBSON: contadores por grupo 1..10, separando Programada vs Urgencia (tipo_atencion).
    Devuelve una tabla simple con grupos en filas y dos columnas (Programada, Urgencia).
    """
    ws = wb.create_sheet("ROBSON")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    headers = ["Grupo Robson", "Descripción (resumen)", "Programada", "Urgencia", "Total"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = bold
        ws.cell(row=1, column=c).alignment = center

    descriptions = {
        1: "Nulípara, único, cefálica, >=37, espontáneo (vaginal)",
        2: "Nulípara, único, cefálica, >=37, inducción o cesárea",
        3: "Multípara sin cesárea previa, único, cefálica, >=37, espontáneo",
        4: "Multípara sin cesárea previa, único, cefálica, >=37, inducción/cesárea",
        5: "Multípara con ≥1 cesárea previa, único, cefálica, ≥37",
        6: "Nulípara, único, podálica",
        7: "Multípara, único, podálica",
        8: "Embarazo múltiple",
        9: "Presentación transversa/oblicua",
        10: "Único, cefálica, <37 semanas",
    }

    thin = Side(border_style="thin", color="000000")
    row = 2
    totals = {"programada": 0, "urgencia": 0, "total": 0}
    for group in range(1, 11):
        # contar programada y urgencia para el grupo
        q_group = [p for p in partidas_qs]  # convert qs to list once outside if needed, but keep simple
        # we'll iterate database-side for correctness: build filter with comprehension not optimal; do python loop
        prog = 0
        urg = 0
        for p in partidas_qs:
            g = robson_group_for_parto(p)
            if g == group:
                if p.tipo_atencion == "programada":
                    prog += 1
                else:
                    urg += 1

        total_g = prog + urg
        totals["programada"] += prog
        totals["urgencia"] += urg
        totals["total"] += total_g

        ws.cell(row=row, column=1, value=f"Grupo {group}")
        ws.cell(row=row, column=2, value=descriptions.get(group, ""))
        ws.cell(row=row, column=3, value=prog)
        ws.cell(row=row, column=4, value=urg)
        ws.cell(row=row, column=5, value=total_g)
        # formato
        ws.cell(row=row, column=1).font = bold
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    # Totales al final
    ws.cell(row=row, column=2, value="Totales").font = bold
    ws.cell(row=row, column=3, value=totals["programada"]).font = bold
    ws.cell(row=row, column=4, value=totals["urgencia"]).font = bold
    ws.cell(row=row, column=5, value=totals["total"]).font = bold

    thin = Side(border_style="thin", color="000000")

    for r in range(1, row):   # desde fila 1 hasta última fila escrita
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )

    # ajustar ancho
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12

# --- View pública --- #

def export_reporte_bs22(request):
    # --- Obtener parámetros con los nombres correctos ---
    start = request.GET.get("inicio")
    end = request.GET.get("fin")

    start_date = None
    end_date = None

    # --- Convertir fechas a objetos date ---
    try:
        if start:
            start_date = datetime.strptime(start, "%Y-%m-%d").date()
        if end:
            end_date = datetime.strptime(end, "%Y-%m-%d").date()
    except Exception:
        start_date = None
        end_date = None

    # --- Query base ---
    partos = Parto.objects.select_related("madre", "registrado_por").order_by("id")

    # --- Aplicar filtros si corresponden ---
    if start_date:
        partos = partos.filter(fecha_parto__gte=start_date)
    if end_date:
        partos = partos.filter(fecha_parto__lte=end_date)

    # --- Crear Excel ---
    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    build_rem_sheet(wb, partos)
    build_aps_sheet(wb, start_date=start_date, end_date=end_date)
    build_robson_sheet(wb, partos)

    # --- Registrar en Bitácora SOLO si el usuario está autenticado ---
    if request.user.is_authenticated:
        Bitacora.objects.create(
            usuario=request.user,
            accion="Generación de reporte REM Bs22",
            detalle=f"Reporte Bs22 desde {start or '(sin inicio)'} hasta {end or '(sin fin)'}"
        )

    # --- Enviar archivo ---
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"reporte_bs22_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

    resp = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'

    return resp

# 📘 Excel REM A09 - Egresos
def exportar_rem_a09(request):
    fecha_inicio = request.GET.get("inicio")
    fecha_fin = request.GET.get("fin")

    partos = Parto.objects.all()
    if fecha_inicio and fecha_fin:
        fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        partos = partos.filter(fecha_parto__range=[fecha_inicio, fecha_fin])

    wb = Workbook()
    ws = wb.active
    ws.title = "REM A09 - Egresos"

    ws.append(["Fecha", "Madre", "Tipo de parto", "Edad gestacional", "Complicaciones", "Nacidos vivos", "Registrado por"])

    for parto in partos:
        registrado_por = ""
        if parto.registrado_por:
            registrado_por = parto.registrado_por.nombre  # usa el campo nombre del modelo Usuario

        ws.append([
            parto.fecha_parto,
            f"{parto.madre.nombres} {parto.madre.apellidos}",
            parto.get_tipo_parto_display(),
            parto.edad_gestacional or "",
            "Sí" if parto.complicaciones else "No",
            parto.recien_nacidos.count(),
            registrado_por
        ])

    # Registrar en bitácora
    Bitacora.objects.create(
        usuario=request.user,
        accion="Generación de reporte REM A09",
        detalle=f"Reporte de egresos desde {fecha_inicio} hasta {fecha_fin}"
    )

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="REM_A09_{datetime.now().date()}.xlsx"'
    wb.save(response)
    return response


# 📗 Excel REM A04 - Defunciones
def exportar_rem_a04(request):
    fecha_inicio = request.GET.get("inicio")
    fecha_fin = request.GET.get("fin")

    rn = RecienNacido.objects.filter(fallecido=True)

    if fecha_inicio and fecha_fin:
        fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        rn = rn.filter(parto__fecha_parto__range=[fecha_inicio, fecha_fin])

    wb = Workbook()
    ws = wb.active
    ws.title = "REM A04 - Defunciones"

    ws.append(["Fecha parto", "Madre", "Edad madre", "Comuna", "Sexo RN", "Tipo fallecimiento", "Matrona responsable"])

    for r in rn:
        matrona = ""
        if r.parto.registrado_por:
            matrona = r.parto.registrado_por.nombre  # usa el campo nombre del usuario

        ws.append([
            r.parto.fecha_parto,
            f"{r.parto.madre.nombres} {r.parto.madre.apellidos}",
            r.parto.madre.edad,
            r.parto.madre.comuna,
            r.get_sexo_display(),
            r.get_tipo_fallecimiento_display() if r.tipo_fallecimiento else "",
            matrona
        ])

    # Registrar en bitácora
    Bitacora.objects.create(
        usuario=request.user,
        accion="Generación de reporte REM A04",
        detalle=f"Reporte de defunciones desde {fecha_inicio} hasta {fecha_fin}"
    )

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="REM_A04_{datetime.now().date()}.xlsx"'
    wb.save(response)
    return response